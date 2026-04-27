"""Excel 워크북 빌더. mode별 (single/tabs/flat) 시트 구성."""
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.shared.aggregation import (
    _group_sort_key,
    sort_categories,
    sort_groups,
    sort_variants,
)

INVALID_SHEET = '\\/?*[]:'
CURRENCY_FMT = '"₩"#,##0'
QTY_FMT = '#,##0'


def sanitize_sheet_name(name, used):
    s = name
    for ch in INVALID_SHEET:
        s = s.replace(ch, "_")
    s = s.strip()[:31] or "sheet"
    base = s
    i = 2
    while s in used:
        suf = f"_{i}"
        s = base[: 31 - len(suf)] + suf
        i += 1
    used.add(s)
    return s


def apply_section_formats(ws):
    """5-column layout: 코드/상품명/단가/판매수/매출"""
    widths = {"A": 14, "B": 60, "C": 13, "D": 10, "E": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    data_font = Font(size=10)
    for r in range(1, ws.max_row + 1):
        a1 = ws.cell(r, 1).value
        is_header = (a1 == "코드") or (a1 == "합계") or (isinstance(a1, str) and a1.startswith("["))
        if not is_header:
            for col in range(1, 6):
                if ws.cell(r, col).font.bold:
                    continue
                ws.cell(r, col).font = data_font
        ws.cell(r, 3).number_format = CURRENCY_FMT
        ws.cell(r, 4).number_format = QTY_FMT
        ws.cell(r, 5).number_format = CURRENCY_FMT


def apply_flat_formats(ws):
    """6-column layout: 카테고리/코드/상품명/단가/판매수/매출"""
    widths = {"A": 30, "B": 14, "C": 60, "D": 13, "E": 10, "F": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    data_font = Font(size=10)
    for r in range(1, ws.max_row + 1):
        a1 = ws.cell(r, 1).value
        is_header = (a1 == "카테고리") or (a1 == "합계")
        if not is_header:
            for col in range(1, 7):
                if ws.cell(r, col).font.bold:
                    continue
                ws.cell(r, col).font = data_font
        ws.cell(r, 4).number_format = CURRENCY_FMT
        ws.cell(r, 5).number_format = QTY_FMT
        ws.cell(r, 6).number_format = CURRENCY_FMT


def build_workbook(results, mode, start, end, currency, sort_by="rev", sort_dir=-1):
    """results = api_report와 동일한 형식 (category_no/category_name/groups/qty/rev).
    mode: 'single' | 'tabs' | 'flat'
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    col_header_font = Font(bold=True, color="FFF8FAFC")
    col_header_fill = PatternFill(start_color="FF334155", end_color="FF334155", fill_type="solid")
    cat_header_font = Font(bold=True, color="FFF8FAFC", size=13)
    cat_header_fill = PatternFill(start_color="FF1E40AF", end_color="FF1E40AF", fill_type="solid")
    sum_fill = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
    parent_fill = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid")
    center_left = Alignment(vertical="center", horizontal="left")

    def emit_section(ws, cat_label, groups):
        ws.append([cat_label])
        cat_row = ws.max_row
        ws.cell(cat_row, 1).font = cat_header_font
        ws.cell(cat_row, 1).fill = cat_header_fill
        ws.cell(cat_row, 1).alignment = center_left
        ws.row_dimensions[cat_row].height = 24
        ws.merge_cells(start_row=cat_row, start_column=1, end_row=cat_row, end_column=5)
        ws.append(["코드", "상품명", "단가", "판매수", "매출"])
        for c in ws[ws.max_row]:
            c.font = col_header_font
            c.fill = col_header_fill
        cat_qty = 0
        cat_rev = 0
        for g in sort_groups(groups, sort_by, sort_dir):
            cat_qty += g["qty"]
            cat_rev += g["rev"]
            if g["is_multi"]:
                parent_name = f"{g['product_name']} ({len(g['variants'])}개 옵션)"
                ws.append([g["product_code"], parent_name, g["price"], g["qty"], g["rev"]])
                p_row = ws.max_row
                for c in ws[p_row]:
                    c.fill = parent_fill
                    c.font = bold
                for v in sort_variants(g["variants"], g["price"], sort_by, sort_dir):
                    suffix = v["variant_code"][len(g["product_code"]):] if v["variant_code"] and v["variant_code"].startswith(g["product_code"]) else (v["variant_code"] or "")
                    label = "  └ " + (v["option"] or v["variant_code"])
                    ws.append([suffix, label, g["price"], v["qty"], v["rev"]])
                    ws.row_dimensions[ws.max_row].outline_level = 1
            else:
                ws.append([g["product_code"], g["product_name"], g["price"], g["qty"], g["rev"]])
        ws.append(["합계", None, None, cat_qty, cat_rev])
        sum_row = ws.max_row
        for c in ws[sum_row]:
            c.font = bold
            c.fill = sum_fill
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
        ws.cell(sum_row, 1).alignment = center_left
        ws.append([])

    if mode == "tabs":
        sorted_results = results
        used = set()
        for r in sorted_results:
            title = sanitize_sheet_name(f"{r['category_no']}_{r['category_name']}", used)
            ws = wb.create_sheet(title=title)
            ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
            ws.append([])
            cat_label = f"[{r['category_no']}] {r['category_name']}"
            emit_section(ws, cat_label, r["groups"])
            ws.freeze_panes = "A4"
            apply_section_formats(ws)
    elif mode == "flat":
        ws = wb.create_sheet(title="전체")
        ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
        ws.append([])
        ws.append(["카테고리", "코드", "상품명", "단가", "판매수", "매출"])
        for c in ws[ws.max_row]:
            c.font = col_header_font
            c.fill = col_header_fill
        all_groups = []
        grand_q = grand_r = 0
        for r in results:
            cat_label = f"[{r['category_no']}] {r['category_name']}"
            for g in r["groups"]:
                all_groups.append((cat_label, g))
                grand_q += g["qty"]
                grand_r += g["rev"]
        all_groups.sort(key=lambda x: _group_sort_key(x[1], sort_by), reverse=(sort_dir == -1))
        for cat_label, g in all_groups:
            if g["is_multi"]:
                parent_name = f"{g['product_name']} ({len(g['variants'])}개 옵션)"
                ws.append([cat_label, g["product_code"], parent_name, g["price"], g["qty"], g["rev"]])
                p_row = ws.max_row
                for c in ws[p_row]:
                    c.fill = parent_fill
                    c.font = bold
                for v in sort_variants(g["variants"], g["price"], sort_by, sort_dir):
                    suffix = v["variant_code"][len(g["product_code"]):] if v["variant_code"] and v["variant_code"].startswith(g["product_code"]) else (v["variant_code"] or "")
                    label = "  └ " + (v["option"] or v["variant_code"])
                    ws.append([cat_label, suffix, label, g["price"], v["qty"], v["rev"]])
                    ws.row_dimensions[ws.max_row].outline_level = 1
            else:
                ws.append([cat_label, g["product_code"], g["product_name"], g["price"], g["qty"], g["rev"]])
        ws.append(["합계", None, None, None, grand_q, grand_r])
        sum_row = ws.max_row
        for c in ws[sum_row]:
            c.font = bold
            c.fill = sum_fill
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
        ws.cell(sum_row, 1).alignment = center_left
        ws.freeze_panes = "A4"
        apply_flat_formats(ws)
    else:
        ws = wb.create_sheet(title="합산")
        ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
        ws.append([])
        sorted_results = results
        for r in sorted_results:
            cat_label = f"[{r['category_no']}] {r['category_name']}"
            emit_section(ws, cat_label, r["groups"])
        ws.freeze_panes = "A3"
        apply_section_formats(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
