"""
cafe24 카테고리별 판매 집계 → Excel.

사용 예:
  py python/cafe24/tests/sales_report.py --category 24 --start 2026-04-01 --end 2026-04-30
  py python/cafe24/tests/sales_report.py --start 2026-04-01 --end 2026-04-30        # depth=1 전체
  py python/cafe24/tests/sales_report.py --category 24                              # 기간 생략 시 지난달
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import argparse
import calendar
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill

from cafe24_auth import get_access_token, MALL_ID

BASE = f"https://{MALL_ID}.cafe24api.com/api/v2/admin"
INVALID_SHEET_CHARS = '\\/?*[]:'


def headers():
    return {"Authorization": f"Bearer {get_access_token()}"}


def fetch_categories():
    res = requests.get(f"{BASE}/categories", headers=headers(), params={"limit": 100})
    res.raise_for_status()
    return res.json().get("categories", [])


def fetch_products_by_category(cat_no):
    products = {}
    offset = 0
    while True:
        res = requests.get(
            f"{BASE}/products", headers=headers(),
            params={"category": cat_no, "limit": 100, "offset": offset, "embed": "variants"},
        )
        res.raise_for_status()
        items = res.json().get("products", [])
        if not items:
            break
        for p in items:
            vs = []
            for v in (p.get("variants") or []):
                opts = v.get("options") or []
                opt_str = ", ".join(
                    f"{o.get('name')}={o.get('value')}"
                    for o in opts if o.get("value")
                ) if opts else ""
                vs.append({"vcode": v.get("variant_code"), "opt": opt_str})
            products[p["product_no"]] = {
                "code": p["product_code"],
                "name": p["product_name"],
                "price": float(p.get("price") or 0),
                "variants": vs,
            }
        if len(items) < 100:
            break
        offset += 100
    return products


def fetch_orders(start_date, end_date):
    orders = []
    offset = 0
    while True:
        res = requests.get(
            f"{BASE}/orders", headers=headers(),
            params={
                "start_date": start_date, "end_date": end_date,
                "embed": "items", "limit": 100, "offset": offset,
            },
        )
        res.raise_for_status()
        page = res.json().get("orders", [])
        if not page:
            break
        orders.extend(page)
        print(f"    ... {len(orders)}건 누적", end="\r", flush=True)
        if len(page) < 100:
            break
        offset += 100
    print(" " * 40, end="\r")
    return orders


def aggregate(products, orders):
    vmap = {}
    for pn, info in products.items():
        multi = len(info["variants"]) > 1
        for v in info["variants"]:
            vc = v["vcode"]
            vmap[vc] = {
                "display_code": vc if multi else info["code"],
                "name": info["name"] + (f" [{v['opt']}]" if multi and v["opt"] else ""),
                "price": info["price"],
                "qty": 0,
                "rev": 0.0,
            }
    for o in orders:
        for it in (o.get("items") or []):
            vc = it.get("variant_code")
            if vc in vmap:
                qty = (it.get("quantity") or 0) - (it.get("claim_quantity") or 0)
                price = float(it.get("product_price") or 0)
                vmap[vc]["qty"] += qty
                vmap[vc]["rev"] += qty * price
    return vmap


def sanitize_sheet_name(name, used):
    s = name
    for ch in INVALID_SHEET_CHARS:
        s = s.replace(ch, "_")
    s = s.strip()[:31] or "sheet"
    base = s
    i = 2
    while s in used:
        suffix = f"_{i}"
        s = (base[: 31 - len(suffix)]) + suffix
        i += 1
    used.add(s)
    return s


def write_excel(out_path, by_category, start, end):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")
    sum_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    used_names = set()
    for cat_label, vmap in by_category.items():
        ws = wb.create_sheet(title=sanitize_sheet_name(cat_label, used_names))
        ws.append([f"기간: {start} ~ {end}", "", "", "", ""])
        ws.append(["코드", "상품명", "단가", "판매수", "매출"])
        for c in ws[2]:
            c.font = bold
            c.fill = header_fill
        rows = sorted(vmap.values(), key=lambda r: -r["rev"])
        total_qty = total_rev = 0
        for r in rows:
            ws.append([r["display_code"], r["name"], r["price"], r["qty"], r["rev"]])
            total_qty += r["qty"]
            total_rev += r["rev"]
        ws.append(["합계", "", "", total_qty, total_rev])
        for c in ws[ws.max_row]:
            c.font = bold
            c.fill = sum_fill
        for col in ("A", "B", "C", "D", "E"):
            max_len = max((len(str(c.value or "")) for c in ws[col]), default=10)
            ws.column_dimensions[col].width = min(max_len + 2, 60)
        ws.freeze_panes = "A3"
    wb.save(out_path)


def default_period():
    today = datetime.today()
    y, m = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
    last_day = calendar.monthrange(y, m)[1]
    return f"{y}-{m:02d}-01", f"{y}-{m:02d}-{last_day:02d}"


def parse_args():
    s, e = default_period()
    p = argparse.ArgumentParser(description="cafe24 카테고리별 판매 집계 → Excel")
    p.add_argument("--category", type=int, help="카테고리 번호. 생략 시 depth=1 전체")
    p.add_argument("--start", default=s, help=f"시작일 YYYY-MM-DD (기본 {s})")
    p.add_argument("--end", default=e, help=f"종료일 YYYY-MM-DD (기본 {e})")
    p.add_argument("--out", default=None, help="출력 xlsx 경로 (기본: d:/sales_report_<period>.xlsx)")
    return p.parse_args()


def main():
    args = parse_args()
    print(f"[1/4] 카테고리 목록 조회")
    cats = fetch_categories()
    print(f"  → {len(cats)}개")

    if args.category is not None:
        target = [c for c in cats if c.get("category_no") == args.category]
        if not target:
            print(f"카테고리 {args.category} 없음", file=sys.stderr)
            sys.exit(1)
    else:
        target = [c for c in cats if c.get("category_depth") == 1]
        print(f"  → 처리 대상(depth=1): {len(target)}개")

    print(f"[2/4] 주문 조회 ({args.start} ~ {args.end})")
    orders = fetch_orders(args.start, args.end)
    print(f"  → 총 {len(orders)}건")

    print(f"[3/4] 카테고리별 상품 조회 + 집계")
    by_category = {}
    for i, c in enumerate(target, 1):
        cat_no = c["category_no"]
        cat_name = c["category_name"]
        label = f"{cat_no}_{cat_name}"
        print(f"  ({i}/{len(target)}) [{cat_no}] {cat_name} ...", end=" ", flush=True)
        products = fetch_products_by_category(cat_no)
        vmap = aggregate(products, orders)
        by_category[label] = vmap
        sold = sum(r["qty"] for r in vmap.values())
        rev = sum(r["rev"] for r in vmap.values())
        print(f"상품 {len(products)} / variant {len(vmap)} / 판매수 {sold} / 매출 {rev:,.0f}")

    out = args.out or f"d:/sales_report_{args.start}_{args.end}.xlsx"
    print(f"[4/4] Excel 저장: {out}")
    write_excel(out, by_category, args.start, args.end)
    print("완료")


if __name__ == "__main__":
    main()
