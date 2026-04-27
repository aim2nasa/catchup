"""
aggregation 모듈 regression 테스트.

실행:
    py python/cafe24/tests/test_aggregation.py
"""
import sys
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(ROOT))

from backend.shared.aggregation import (
    aggregate,
    sort_groups,
    sort_variants,
    sort_categories,
)


def make_product(pn, code, name, price, variants):
    """variants = [(variant_code, option), ...]"""
    return (pn, {
        "code": code,
        "name": name,
        "price": float(price),
        "variants": [{"vcode": v[0], "opt": v[1]} for v in variants],
    })


def make_products(*args):
    return dict(args)


def make_order(items):
    """items = [(variant_code, quantity, product_price[, claim_quantity]), ...]"""
    return {"items": [{
        "variant_code": it[0],
        "quantity": it[1],
        "product_price": float(it[2]),
        "claim_quantity": it[3] if len(it) > 3 else 0,
    } for it in items]}


class TestAggregate(unittest.TestCase):

    def test_empty_inputs(self):
        self.assertEqual(aggregate({}, []), [])

    def test_single_variant_no_orders(self):
        products = make_products(make_product(1, "P0001", "Item A", 100, [("P0001000A", "")]))
        groups = aggregate(products, [])
        self.assertEqual(len(groups), 1)
        g = groups[0]
        self.assertFalse(g["is_multi"])
        self.assertEqual(g["product_code"], "P0001")
        self.assertEqual(g["product_name"], "Item A")
        self.assertEqual(g["price"], 100.0)
        self.assertEqual(g["qty"], 0)
        self.assertEqual(g["rev"], 0.0)
        self.assertEqual(len(g["variants"]), 1)
        self.assertEqual(g["variants"][0]["variant_code"], "P0001000A")

    def test_single_variant_one_order(self):
        products = make_products(make_product(1, "P0001", "A", 100, [("P0001000A", "")]))
        orders = [make_order([("P0001000A", 3, 100)])]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 3)
        self.assertEqual(g["rev"], 300.0)
        self.assertEqual(g["variants"][0]["qty"], 3)
        self.assertEqual(g["variants"][0]["rev"], 300.0)

    def test_claim_quantity_subtracted(self):
        products = make_products(make_product(1, "P0001", "X", 100, [("P0001000A", "")]))
        orders = [make_order([("P0001000A", 5, 100, 2)])]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 3)
        self.assertEqual(g["rev"], 300.0)

    def test_multi_variant_aggregation(self):
        products = make_products(make_product(1, "P0002", "Combo", 0, [
            ("P0002000A", "A"), ("P0002000B", "B"), ("P0002000C", "C"),
        ]))
        orders = [
            make_order([("P0002000A", 2, 50)]),
            make_order([("P0002000B", 5, 50), ("P0002000C", 1, 50)]),
            make_order([("P0002000A", 3, 50)]),
        ]
        g = aggregate(products, orders)[0]
        self.assertTrue(g["is_multi"])
        self.assertEqual(g["qty"], 11)
        self.assertEqual(g["rev"], 550.0)
        # parent qty/rev == sum of variants
        self.assertEqual(g["qty"], sum(v["qty"] for v in g["variants"]))
        self.assertEqual(g["rev"], sum(v["rev"] for v in g["variants"]))
        # individual values
        v_map = {v["variant_code"]: v for v in g["variants"]}
        self.assertEqual(v_map["P0002000A"]["qty"], 5)
        self.assertEqual(v_map["P0002000A"]["rev"], 250.0)
        self.assertEqual(v_map["P0002000B"]["qty"], 5)
        self.assertEqual(v_map["P0002000B"]["rev"], 250.0)
        self.assertEqual(v_map["P0002000C"]["qty"], 1)
        self.assertEqual(v_map["P0002000C"]["rev"], 50.0)

    def test_zero_price_bundle(self):
        products = make_products(make_product(1, "PVM", "Bundle", 0, [
            ("PVM000A", "opt1"), ("PVM000B", "opt2"),
        ]))
        orders = [make_order([("PVM000A", 100, 0), ("PVM000B", 50, 0)])]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 150)
        self.assertEqual(g["rev"], 0.0)

    def test_unknown_variant_ignored(self):
        products = make_products(make_product(1, "P0001", "X", 100, [("P0001000A", "")]))
        orders = [make_order([("P0001000A", 2, 100), ("PUNKNOWN000A", 99, 999)])]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 2)
        self.assertEqual(g["rev"], 200.0)

    def test_multiple_products(self):
        products = make_products(
            make_product(1, "P001", "A", 100, [("P001000A", "")]),
            make_product(2, "P002", "B", 200, [("P002000A", "")]),
        )
        orders = [make_order([("P001000A", 3, 100), ("P002000A", 1, 200)])]
        groups = aggregate(products, orders)
        self.assertEqual(len(groups), 2)
        m = {g["product_code"]: g for g in groups}
        self.assertEqual(m["P001"]["qty"], 3)
        self.assertEqual(m["P001"]["rev"], 300.0)
        self.assertEqual(m["P002"]["qty"], 1)
        self.assertEqual(m["P002"]["rev"], 200.0)

    def test_grand_total_consistency(self):
        """카테고리 합계가 사이트 grand와 일치해야 함."""
        products_cat1 = make_products(make_product(1, "P001", "A", 100, [("P001000A", "")]))
        products_cat2 = make_products(make_product(2, "P002", "B", 200, [("P002000A", "")]))
        orders = [
            make_order([("P001000A", 3, 100)]),
            make_order([("P002000A", 2, 200)]),
        ]
        cat1 = aggregate(products_cat1, orders)
        cat2 = aggregate(products_cat2, orders)
        cat1_qty = sum(g["qty"] for g in cat1)
        cat2_qty = sum(g["qty"] for g in cat2)
        cat1_rev = sum(g["rev"] for g in cat1)
        cat2_rev = sum(g["rev"] for g in cat2)
        self.assertEqual(cat1_qty + cat2_qty, 5)
        self.assertEqual(cat1_rev + cat2_rev, 700.0)

    def test_no_double_count_across_orders(self):
        """같은 variant가 여러 주문에 나오면 합산되어야지 중복되면 안 됨."""
        products = make_products(make_product(1, "P001", "X", 100, [("P001000A", "")]))
        orders = [
            make_order([("P001000A", 1, 100)]),
            make_order([("P001000A", 2, 100)]),
            make_order([("P001000A", 4, 100)]),
        ]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 7)
        self.assertEqual(g["rev"], 700.0)

    def test_negative_quantity_after_claim(self):
        """claim이 quantity보다 크면 음수 가능 (cafe24 실데이터에서도 발생)."""
        products = make_products(make_product(1, "P001", "X", 100, [("P001000A", "")]))
        orders = [make_order([("P001000A", 2, 100, 5)])]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], -3)
        self.assertEqual(g["rev"], -300.0)

    def test_missing_variant_code_in_order_skipped(self):
        products = make_products(make_product(1, "P001", "X", 100, [("P001000A", "")]))
        # variant_code=None인 item은 무시되어야 함
        orders = [{"items": [
            {"variant_code": None, "quantity": 99, "product_price": 999, "claim_quantity": 0},
            {"variant_code": "P001000A", "quantity": 2, "product_price": 100, "claim_quantity": 0},
        ]}]
        g = aggregate(products, orders)[0]
        self.assertEqual(g["qty"], 2)
        self.assertEqual(g["rev"], 200.0)


class TestSortGroups(unittest.TestCase):

    def setUp(self):
        self.groups = [
            {"product_code": "P00003", "product_name": "Charlie", "price": 30, "qty": 10, "rev": 200, "is_multi": False, "variants": []},
            {"product_code": "P00001", "product_name": "Alpha",   "price": 10, "qty": 5,  "rev": 100, "is_multi": False, "variants": []},
            {"product_code": "P00002", "product_name": "Bravo",   "price": 20, "qty": 3,  "rev": 500, "is_multi": False, "variants": []},
        ]

    def test_sort_by_rev_desc(self):
        s = sort_groups(self.groups, "rev", -1)
        self.assertEqual([g["product_code"] for g in s], ["P00002", "P00003", "P00001"])

    def test_sort_by_rev_asc(self):
        s = sort_groups(self.groups, "rev", 1)
        self.assertEqual([g["product_code"] for g in s], ["P00001", "P00003", "P00002"])

    def test_sort_by_qty_desc(self):
        s = sort_groups(self.groups, "qty", -1)
        self.assertEqual([g["product_code"] for g in s], ["P00003", "P00001", "P00002"])

    def test_sort_by_price_asc(self):
        s = sort_groups(self.groups, "price", 1)
        self.assertEqual([g["product_code"] for g in s], ["P00001", "P00002", "P00003"])

    def test_sort_by_code_asc(self):
        s = sort_groups(self.groups, "code", 1)
        self.assertEqual([g["product_code"] for g in s], ["P00001", "P00002", "P00003"])

    def test_sort_by_code_desc(self):
        s = sort_groups(self.groups, "code", -1)
        self.assertEqual([g["product_code"] for g in s], ["P00003", "P00002", "P00001"])

    def test_sort_by_name_asc(self):
        s = sort_groups(self.groups, "name", 1)
        self.assertEqual([g["product_code"] for g in s], ["P00001", "P00002", "P00003"])

    def test_sort_does_not_mutate_input(self):
        before = list(self.groups)
        sort_groups(self.groups, "rev", -1)
        self.assertEqual(self.groups, before)


class TestSortVariants(unittest.TestCase):

    def test_sort_variants_by_qty(self):
        vars = [
            {"variant_code": "X1", "option": "a", "qty": 5, "rev": 100},
            {"variant_code": "X2", "option": "b", "qty": 1, "rev": 50},
            {"variant_code": "X3", "option": "c", "qty": 9, "rev": 200},
        ]
        s = sort_variants(vars, parent_price=100, sort_by="qty", sort_dir=-1)
        self.assertEqual([v["variant_code"] for v in s], ["X3", "X1", "X2"])

    def test_sort_variants_by_option_name(self):
        vars = [
            {"variant_code": "X1", "option": "라벤더", "qty": 5, "rev": 100},
            {"variant_code": "X2", "option": "가벼운향",  "qty": 1, "rev": 50},
            {"variant_code": "X3", "option": "딸기",   "qty": 9, "rev": 200},
        ]
        s = sort_variants(vars, parent_price=100, sort_by="name", sort_dir=1)
        self.assertEqual([v["option"] for v in s], ["가벼운향", "딸기", "라벤더"])


class TestSortCategories(unittest.TestCase):

    def setUp(self):
        self.results = [
            {"category_no": 30, "category_name": "다", "qty": 5,  "rev": 100, "groups": []},
            {"category_no": 10, "category_name": "가", "qty": 9,  "rev": 500, "groups": []},
            {"category_no": 20, "category_name": "나", "qty": 12, "rev": 200, "groups": []},
        ]

    def test_by_rev_desc(self):
        s = sort_categories(self.results, "rev", -1)
        self.assertEqual([r["category_no"] for r in s], [10, 20, 30])

    def test_by_qty_desc(self):
        # cat 20 qty=12, cat 10 qty=9, cat 30 qty=5
        s = sort_categories(self.results, "qty", -1)
        self.assertEqual([r["category_no"] for r in s], [20, 10, 30])

    def test_by_cat_no_asc(self):
        s = sort_categories(self.results, "code", 1)
        self.assertEqual([r["category_no"] for r in s], [10, 20, 30])

    def test_by_name_asc(self):
        s = sort_categories(self.results, "name", 1)
        self.assertEqual([r["category_name"] for r in s], ["가", "나", "다"])


class TestExcelParityAcceptance(unittest.TestCase):
    """downloaded Excel과 aggregate 결과 1원/1개 오차 없음 검증.

    이 테스트는 실제 파일이 있을 때만 실행. CI 환경에선 자동 skip.
    """

    EXCEL = Path(r"D:\Users\rossi\Downloads\sales_report_2026-02-28_2026-03-30 (1).xlsx")

    @classmethod
    def setUpClass(cls):
        if not cls.EXCEL.exists():
            raise unittest.SkipTest(f"Excel 파일 없음: {cls.EXCEL}")
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise unittest.SkipTest("openpyxl 미설치")

    def test_excel_parent_only_sum_consistency(self):
        """Excel의 parent-only 합 = Excel의 합계 row 값 (자체 검증)."""
        import openpyxl
        wb = openpyxl.load_workbook(self.EXCEL, data_only=True)
        ws = wb[wb.sheetnames[0]]
        parent_q = parent_r = 0
        total_row_q = total_row_r = None
        for r in range(3, ws.max_row + 1):
            cat = ws.cell(r, 1).value
            if cat == "합계":
                total_row_q = ws.cell(r, 5).value
                total_row_r = ws.cell(r, 6).value
                continue
            name = str(ws.cell(r, 3).value or "")
            if not name.strip().startswith("└"):
                parent_q += ws.cell(r, 5).value or 0
                parent_r += ws.cell(r, 6).value or 0
        self.assertIsNotNone(total_row_q)
        self.assertEqual(parent_q, total_row_q)
        self.assertEqual(parent_r, total_row_r)


if __name__ == "__main__":
    unittest.main(verbosity=2)
