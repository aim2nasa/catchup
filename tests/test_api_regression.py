"""
catchup web 백엔드 회귀 테스트.

cafe24 외부 API는 tests/fixtures/*.json 으로 mock하여, 향후 리팩토링/구조변경
이후에도 동일한 결과(JSON shape, 합계, Excel 구조)가 유지되는지 검증.

실행:
    py -m pytest tests/test_api_regression.py -v
또는:
    py tests/test_api_regression.py
"""
import io
import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

ROOT = Path(__file__).resolve().parent.parent
FIX = Path(__file__).resolve().parent / "fixtures"
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "python" / "cafe24" / "tests"))

# cafe24_auth.get_access_token을 stub (network/token 없이 import 가능하도록)
import cafe24_auth as _cafe24_auth  # type: ignore
_cafe24_auth.get_access_token = lambda: "FAKE_TOKEN"

from backend.main import app  # noqa: E402

from fastapi.testclient import TestClient  # noqa: E402

import openpyxl  # noqa: E402


def _load_fixture(name: str):
    return json.loads((FIX / name).read_text(encoding="utf-8"))


class FakeResp:
    def __init__(self, payload, status=200):
        self._p = payload
        self.status_code = status
        self.text = ""

    def json(self):
        return self._p

    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError(f"HTTP {self.status_code}")


def fake_requests_get(url, headers=None, params=None, **kw):
    """cafe24 API 4종 (categories / products / orders) mock.

    - categories: cafe24_categories.json (93개)
    - products?category=24: cafe24_products_cat24.json (26개)
    - products?category=<other>: 빈 리스트
    - orders: cafe24_orders_small.json (5개) 한 페이지로 끝
    """
    params = params or {}
    if url.endswith("/categories"):
        return FakeResp(_load_fixture("cafe24_categories.json"))
    if url.endswith("/products"):
        cat = params.get("category")
        offset = int(params.get("offset", 0))
        if offset > 0:
            return FakeResp({"products": []})
        if cat == 24:
            return FakeResp(_load_fixture("cafe24_products_cat24.json"))
        return FakeResp({"products": []})
    if url.endswith("/orders"):
        offset = int(params.get("offset", 0))
        if offset > 0:
            return FakeResp({"orders": []})
        return FakeResp(_load_fixture("cafe24_orders_small.json"))
    raise AssertionError(f"unmocked url: {url}")


class _Patched(unittest.TestCase):
    """공통 setUp/tearDown — requests.get을 글로벌하게 mock."""

    def setUp(self):
        self._patcher = patch("requests.get", side_effect=fake_requests_get)
        self._patcher.start()
        self.client = TestClient(app)

    def tearDown(self):
        self._patcher.stop()


class TestVersionEndpoint(_Patched):
    def test_version_shape(self):
        r = self.client.get("/api/version")
        self.assertEqual(r.status_code, 200)
        d = r.json()
        self.assertIn("version", d)
        self.assertIn("started_at", d)
        self.assertIsInstance(d["version"], str)


class TestAdminRestartEndpoint(_Patched):
    def test_rejects_backend_standalone_restart_without_supervisor(self):
        with patch("backend.domains.catalog.routes.shutil.which", return_value=None), patch.dict(
            os.environ,
            {"CATCHUP_DEV_RESTART_FILE": ""},
        ):
            r = self.client.post("/api/admin/restart")

        self.assertEqual(r.status_code, 409)
        data = r.json()
        self.assertFalse(data["ok"])
        self.assertIn("안전한 서버 재시작", data["error"])
        self.assertIn("npm run dev", data["error"])

    def test_dev_supervisor_restart_writes_request_file_without_exiting_backend(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            restart_file = Path(temp_dir) / "restart-request.json"
            with patch("backend.domains.catalog.routes.shutil.which", return_value=None), patch.dict(
                os.environ,
                {"CATCHUP_DEV_RESTART_FILE": str(restart_file)},
            ):
                r = self.client.post("/api/admin/restart")

            self.assertEqual(r.status_code, 200)
            data = r.json()
            self.assertTrue(data["ok"])
            self.assertEqual(data["mode"], "dev-supervisor")
            self.assertTrue(restart_file.exists())

            payload = json.loads(restart_file.read_text(encoding="utf-8"))
            self.assertEqual(payload["mode"], "dev-supervisor")
            self.assertIn("requested_at", payload)
            self.assertIn("version", payload)


class TestCategoriesEndpoint(_Patched):
    def test_returns_flat_list(self):
        r = self.client.get("/api/categories")
        self.assertEqual(r.status_code, 200)
        data = r.json()
        self.assertIsInstance(data, list)
        self.assertGreater(len(data), 0)

    def test_keys_consistent(self):
        r = self.client.get("/api/categories")
        for c in r.json():
            self.assertSetEqual(set(c.keys()), {"no", "name", "depth", "parent"})

    def test_includes_known_cats(self):
        r = self.client.get("/api/categories")
        nos = {c["no"] for c in r.json()}
        # 주요 depth=1 카테고리 — 스냅샷 기준
        for known in (24, 25, 26):
            self.assertIn(known, nos)


class TestReportSSE(_Patched):
    def _collect_events(self, params):
        with self.client.stream("GET", "/api/report", params=params) as r:
            self.assertEqual(r.status_code, 200)
            events = []
            for line in r.iter_lines():
                if line.startswith("data: "):
                    events.append(json.loads(line[6:]))
            return events

    def test_sse_event_sequence(self):
        evs = self._collect_events({
            "start": "2026-04-25", "end": "2026-04-26", "categories": "24",
        })
        types = [e["type"] for e in evs]
        self.assertIn("data", types)
        self.assertEqual(types[-1], "done")
        # data 이벤트는 정확히 1회
        self.assertEqual(types.count("data"), 1)

    def test_data_payload_shape(self):
        evs = self._collect_events({
            "start": "2026-04-25", "end": "2026-04-26", "categories": "24",
        })
        data = next(e for e in evs if e["type"] == "data")
        self.assertIn("results", data)
        self.assertIn("grand", data)
        self.assertIn("start", data)
        self.assertIn("end", data)
        g = data["grand"]
        self.assertSetEqual(set(g.keys()), {"qty", "rev", "currency", "order_count"})
        # currency: orders fixture에 currency 필드가 있으면 반영, 없으면 KRW
        self.assertIn(g["currency"], {"KRW"})

    def test_grand_equals_sum_of_categories(self):
        evs = self._collect_events({
            "start": "2026-04-25", "end": "2026-04-26", "categories": "24",
        })
        data = next(e for e in evs if e["type"] == "data")
        sum_q = sum(r["qty"] for r in data["results"])
        sum_r = sum(r["rev"] for r in data["results"])
        self.assertEqual(data["grand"]["qty"], sum_q)
        self.assertAlmostEqual(data["grand"]["rev"], sum_r, places=4)

    def test_results_in_user_order(self):
        # 카테고리 입력 순서대로 결과가 나와야 함 (parse_categories 보존)
        evs = self._collect_events({
            "start": "2026-04-25", "end": "2026-04-26", "categories": "26,24,25",
        })
        data = next(e for e in evs if e["type"] == "data")
        nos = [r["category_no"] for r in data["results"]]
        # 빈 카테고리(25)도 포함되며 사용자 순서 유지
        self.assertEqual(nos, [26, 24, 25])

    def test_group_internal_consistency(self):
        evs = self._collect_events({
            "start": "2026-04-25", "end": "2026-04-26", "categories": "24",
        })
        data = next(e for e in evs if e["type"] == "data")
        # multi-variant: parent.qty == sum(variants.qty), parent.rev == sum(variants.rev)
        for cat in data["results"]:
            for g in cat["groups"]:
                sv_q = sum(v["qty"] for v in g["variants"])
                sv_r = sum(v["rev"] for v in g["variants"])
                self.assertEqual(g["qty"], sv_q, msg=f"{g['product_code']} qty 불일치")
                self.assertAlmostEqual(g["rev"], sv_r, places=4, msg=f"{g['product_code']} rev 불일치")


class TestProductsReportRequestSSE(_Patched):
    def _collect_stream_events(self, request_id):
        with self.client.stream("GET", f"/api/products-report-stream/{request_id}") as r:
            self.assertEqual(r.status_code, 200)
            events = []
            for line in r.iter_lines():
                if line.startswith("data: "):
                    events.append(json.loads(line[6:]))
            return events

    def test_product_codes_report_uses_short_registered_stream_url(self):
        codes = [f"P{i:07d}" for i in range(250)]
        r = self.client.post("/api/products-report-requests", json={
            "start": "2026-04-25",
            "end": "2026-04-26",
            "codes": codes,
        })
        self.assertEqual(r.status_code, 200)
        payload = r.json()
        self.assertRegex(payload["request_id"], r"^[0-9a-f]{32}$")

        evs = self._collect_stream_events(payload["request_id"])
        types = [e["type"] for e in evs]
        self.assertIn("data", types)
        self.assertEqual(types[-1], "done")
        progress = [e.get("msg", "") for e in evs if e["type"] == "progress"]
        self.assertTrue(any("250개 코드" in msg for msg in progress))

    def test_product_codes_report_request_validates_required_codes(self):
        r = self.client.post("/api/products-report-requests", json={
            "start": "2026-04-25",
            "end": "2026-04-26",
            "codes": [],
        })
        self.assertEqual(r.status_code, 400)

    def test_unknown_product_codes_stream_request_returns_404(self):
        r = self.client.get("/api/products-report-stream/not-found")
        self.assertEqual(r.status_code, 404)


class TestExcelDownload(_Patched):
    def _download(self, mode):
        r = self.client.get("/api/excel", params={
            "start": "2026-04-25", "end": "2026-04-26",
            "categories": "24", "mode": mode,
        })
        self.assertEqual(r.status_code, 200)
        self.assertIn("application/vnd.openxmlformats", r.headers["content-type"])
        return openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    def test_tabs_mode_produces_per_cat_sheet(self):
        wb = self._download("tabs")
        # 시트 1개 (cat 24만)
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb[wb.sheetnames[0]]
        # 카테고리 헤더 행 + 컬럼 헤더 행 + 데이터 + 합계
        first_row_a = ws.cell(1, 1).value
        self.assertTrue(str(first_row_a).startswith("기간:"))
        # "코드" 컬럼 헤더 행이 어딘가 있어야 함
        col_headers = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        self.assertIn("코드", col_headers)
        self.assertIn("합계", col_headers)

    def test_single_mode_one_sheet(self):
        wb = self._download("single")
        self.assertEqual(len(wb.sheetnames), 1)
        self.assertEqual(wb.sheetnames[0], "합산")

    def test_flat_mode_has_category_column(self):
        wb = self._download("flat")
        self.assertEqual(wb.sheetnames[0], "전체")
        ws = wb[wb.sheetnames[0]]
        # row 3 = 컬럼 헤더 (기간:, 빈줄 다음)
        # "카테고리" 헤더 셀 검색
        cells_a = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        self.assertIn("카테고리", cells_a)

    def test_no_autofilter(self):
        # 정렬 기능(autofilter)은 의도적으로 제거됨
        wb = self._download("single")
        ws = wb[wb.sheetnames[0]]
        self.assertIsNone(ws.auto_filter.ref)

    def test_currency_format_applied(self):
        wb = self._download("tabs")
        ws = wb[wb.sheetnames[0]]
        # 데이터 행에서 매출 셀의 number_format 확인
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, 5).value
            if isinstance(v, (int, float)):
                self.assertEqual(ws.cell(r, 5).number_format, '"₩"#,##0')
                return
        self.fail("매출 셀을 찾지 못함")


if __name__ == "__main__":
    unittest.main(verbosity=2)
