import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
import requests

EXCEL = Path(r'D:\catchup\docs\상품출고&순이익(2026)_NEW1_cafe24_1to1_checked_with_U_sales_v6.xlsx')
SHEET = '상품코드_4월-NEW1'
START = '2026-04-01'
END = '2026-04-30'

L_GROUPS = [
    ('500g', ['P00000HT', 'P00000BV', 'P00000CB', 'P00000BX', 'P00000XE', 'P0000BIF', 'P0000BLD', 'P0000BMJ', 'P0000BMI'], True),
    ('cupb', ['P00000ZB'], False),
    ('1kg', ['P00000UH', 'P00000TI', 'P00000BY', 'P00000BZ', 'P00000CH', 'P00000CG', 'P00000CA', 'P00000BW', 'P00000CI', 'P00000CE', 'P00000KH', 'P00000CD', 'P00000CF'], True),
]
L_CODES = [c for _, codes, _ in L_GROUPS for c in codes]

U_BLOCKS = [
    ('P00000QE', ['G', 'H', 'I', 'J', 'K']),
    ('P00000QD', ['CI', 'CJ', 'CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT', 'CU']),
    ('P0000BLR', ['Q', 'R', 'S', 'T', 'U', 'V', 'W']),
    ('P0000BLA', ['J', 'K', 'L', 'M', 'N', 'O', 'P']),
    ('P00000ZC', ['A', 'B', 'C', 'D']),
    ('P00000YZ', ['D', 'E', 'H', 'I']),
    ('P00000VM', ['CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD', 'DE', 'DF', 'DG', 'DH']),
    ('P00000YS', ['A']),
    ('P00000YU', ['B']),
]

U_COLUMNS = [(u, variant) for u, variants in U_BLOCKS for variant in variants]

RULES = [
    ('P00000QE', 'G', 'P00000BV', None, 1),
    ('P00000QE', 'I', 'P00000CB', None, 1),
    ('P00000QE', 'J', 'P00000BX', None, 1),
    ('P00000QE', 'K', 'P00000XE', None, 1),
    ('P00000QD', 'CI', 'P00000CF', None, 1),
    ('P00000QD', 'CJ', 'P00000CE', None, 1),
    ('P00000QD', 'CK', 'P00000KH', None, 1),
    ('P00000QD', 'CL', 'P00000CA', None, 1),
    ('P00000QD', 'CM', 'P00000CG', None, 1),
    ('P00000QD', 'CN', 'P00000BW', None, 1),
    ('P00000QD', 'CO', 'P00000CD', None, 1),
    ('P00000QD', 'CP', 'P00000CH', None, 1),
    ('P00000QD', 'CQ', 'P00000CI', None, 1),
    ('P00000QD', 'CR', 'P00000BY', None, 1),
    ('P00000QD', 'CS', 'P00000BZ', None, 1),
    ('P00000QD', 'CT', 'P00000TI', None, 1),
    ('P00000QD', 'CU', 'P00000UH', None, 1),
    ('P0000BLR', 'Q', 'P0000BIF', 'CI', 1),
    ('P0000BLR', 'R', 'P0000BIF', 'CJ', 1),
    ('P0000BLR', 'S', 'P0000BIF', 'CN', 1),
    ('P0000BLR', 'T', 'P0000BIF', 'CL', 1),
    ('P0000BLR', 'U', 'P0000BIF', 'CM', 1),
    ('P0000BLR', 'V', 'P0000BIF', 'CK', 1),
    ('P0000BLR', 'W', 'P0000BIF', 'CO', 1),
    ('P0000BLA', 'J', 'P0000BIF', 'CI', 1),
    ('P0000BLA', 'K', 'P0000BIF', 'CJ', 1),
    ('P0000BLA', 'L', 'P0000BIF', 'CK', 1),
    ('P0000BLA', 'M', 'P0000BIF', 'CN', 1),
    ('P0000BLA', 'N', 'P0000BIF', 'CL', 1),
    ('P0000BLA', 'O', 'P0000BIF', 'CM', 1),
    ('P0000BLA', 'P', 'P0000BIF', 'CO', 1),
    ('P00000ZC', 'A', 'P00000ZB', None, 1),
    ('P00000ZC', 'B', 'P00000ZB', None, 1),
    ('P00000ZC', 'C', 'P00000ZB', None, 1),
    ('P00000ZC', 'D', 'P00000ZB', None, 1),
    ('P00000YZ', 'D', 'P00000ZB', None, 1),
    ('P00000YZ', 'E', 'P00000ZB', None, 1),
    ('P00000YZ', 'H', 'P00000ZB', None, 1),
    ('P00000YZ', 'I', 'P00000ZB', None, 1),
    ('P00000VM', 'CW', 'P00000TI', None, 1),
    ('P00000VM', 'CX', 'P00000CD', None, 1),
    ('P00000VM', 'CY', 'P00000CA', None, 1),
    ('P00000VM', 'CZ', 'P00000CE', None, 1),
    ('P00000VM', 'DA', 'P00000CF', None, 1),
    ('P00000VM', 'DB', 'P00000BY', None, 1),
    ('P00000VM', 'DC', 'P00000BW', None, 1),
    ('P00000VM', 'DD', 'P00000BZ', None, 1),
    ('P00000VM', 'DE', 'P00000CI', None, 1),
    ('P00000VM', 'DF', 'P00000CG', None, 1),
    ('P00000VM', 'DG', 'P00000CH', None, 1),
    ('P00000VM', 'DH', 'P00000UH', None, 1),
    ('P00000YS', 'A', 'P00000HT', None, 1),
]

EXCLUDED_U_PRODUCTS = {'P00000YU'}

RULE_BY_KEY = defaultdict(list)
for rule in RULES:
    key = f"{rule[0]}|{rule[1]}"
    RULE_BY_KEY[key].append(rule)

rule_u_idx = {}
variant_counters = {}
for u_prod, u_var in U_COLUMNS:
    rule_u_idx[(u_prod, u_var)] = variant_counters.get(u_prod, 0)
    variant_counters[u_prod] = variant_counters.get(u_prod, 0) + 1


def norm_suffix(product_code: str, variant_code: str) -> str:
    if not variant_code:
        return ''
    raw = str(variant_code)
    if raw.startswith(product_code):
        raw = raw[len(product_code):]
    raw = raw.lstrip('0')
    return (raw or raw).upper()


def get_rule_match_qty(candidates, target_l_product, target_l_variant, target_l_variant_index, rule_u_variant_index, qty, has_l_variants):
    if not candidates:
        return 0
    total_qty = 0
    for _u_prod, _u_var, l_product, l_variant, ratio in candidates:
        if l_product != target_l_product:
            continue
        if l_variant is not None:
            if not target_l_variant:
                continue
            if l_variant != target_l_variant:
                continue
            total_qty += qty * ratio
            continue

        if target_l_variant is None:
            if not has_l_variants:
                total_qty += qty * ratio
            continue

        if target_l_variant_index is not None and rule_u_variant_index == target_l_variant_index:
            total_qty += qty * ratio
    return total_qty


def get_api_data():
    url = 'http://localhost:8000/api/products-report'
    params = {
        'start': START,
        'end': END,
        'codes': ','.join(sorted(set(L_CODES + [u for u, _ in U_COLUMNS]))),
    }
    r = requests.get(url, params=params, stream=True, timeout=120)
    r.raise_for_status()

    payload = None
    for line in r.iter_lines(decode_unicode=True):
        if not line or not line.startswith('data:'):
            continue
        body = line.split('data:', 1)[1].strip()
        if not body or body == '[DONE]':
            continue
        data = json.loads(body)
        if data.get('type') == 'data':
            payload = data
            break
    if payload is None:
        raise RuntimeError('No data event received')
    return {g['product_code']: g for g in payload['results'][0]['groups']}


def col_to_colidx(ws):
    # Map sheet columns 5~58 to U_COLUMNS index
    mapping = {}

    def merged_value(row, col):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(rng.min_row, rng.min_col).value
        return None

    for col in range(5, 59):
        prod = ws.cell(6, col).value
        if prod is None:
            prod = merged_value(6, col)
        var = ws.cell(8, col).value
        if prod is None or var is None:
            continue
        prod = str(prod).strip()
        var = str(var).replace(' ', '').strip().upper()
        for idx, (u_prod, u_var) in enumerate(U_COLUMNS):
            if u_prod == prod and u_var == var:
                mapping[col] = idx
                break

    return mapping


def u_variant_qty(by_code, u_prod, u_var):
    if u_prod in EXCLUDED_U_PRODUCTS:
        return None
    g = by_code.get(u_prod)
    if not g:
        return 0
    for v in g.get('variants', []):
        if norm_suffix(u_prod, v['variant_code']) == u_var:
            return v['qty']
    return 0


def row_expected(by_code, row):
    code = row['code']
    variant = row['variant']
    g = by_code.get(code)
    has_l_variants = bool(g and g.get('variants') and len(g.get('variants')) > 1)
    direct = 0
    l_variant_index = None

    variants = g.get('variants', []) if g else []
    if variant:
        if has_l_variants:
            for i, vv in enumerate(variants):
                if norm_suffix(code, vv['variant_code']) == variant:
                    l_variant_index = i
                    direct = vv['qty']
                    break
        else:
            # 단일 옵션 상품은 엑셀 셀의 A/B 같은 품목코드 표기를 부모 행 배치로 해석한다.
            l_variant_index = None
            direct = g['qty'] if g else 0
        if l_variant_index is None and variant == 'A' and has_l_variants:
            l_variant_index = 0
    elif g:
        direct = g['qty']

    mapping = {idx: 0 for idx in range(len(U_COLUMNS))}

    for idx, (u_prod, u_var) in enumerate(U_COLUMNS):
        if u_prod in EXCLUDED_U_PRODUCTS:
            continue

        u_qty = u_variant_qty(by_code, u_prod, u_var)
        if u_qty is None:
            continue

        candidates = RULE_BY_KEY.get(f'{u_prod}|{u_var}')
        matched = get_rule_match_qty(
            candidates,
            code,
            variant if has_l_variants else None,
            l_variant_index,
            rule_u_idx.get((u_prod, u_var)),
            u_qty,
            has_l_variants,
        )
        if matched:
            mapping[idx] = matched

    return direct, mapping


def parse_rows(ws):
    rows = []
    current_code = None

    for r in range(10, 48):
        code_v = ws.cell(r, 1).value
        var_v = ws.cell(r, 2).value
        direct = ws.cell(r, 4).value
        try:
            direct_num = int(direct) if direct is not None else None
        except (TypeError, ValueError):
            direct_num = None

        cols = {c: ws.cell(r, c).value for c in range(5, 59)}

        code = str(code_v).strip() if code_v is not None else None
        variant = str(var_v).strip().upper() if var_v is not None else None

        if code in ('500g총합계', '1kg총합계', '합계', '2kg총합계'):
            current_code = None
            rows.append({'row': r, 'kind': 'subtotal', 'label': code, 'code': code, 'variant': None, 'direct': direct_num or 0, 'cols': cols})
            continue

        if code:
            current_code = code
            if code in L_CODES:
                rows.append({'row': r, 'kind': 'data', 'code': code, 'variant': variant, 'direct': direct_num or 0, 'cols': cols})
            else:
                # unknown product code in sheet
                pass
            continue

        if not code and variant and current_code:
            rows.append({'row': r, 'kind': 'data', 'code': current_code, 'variant': variant, 'direct': direct_num or 0, 'cols': cols})
            continue

    return rows


def excel_col(idx: int) -> str:
    n = idx
    s = ''
    while n > 0:
        n -= 1
        s = chr(65 + (n % 26)) + s
        n //= 26
    return s


def compare():
    ws = load_workbook(EXCEL, data_only=True)[SHEET]
    by_code = get_api_data()
    col_map = col_to_colidx(ws)

    rows = parse_rows(ws)
    mismatches = []

    # U direct row (row9)
    for col, idx in sorted(col_map.items()):
        u_prod, u_var = U_COLUMNS[idx]
        if u_prod in EXCLUDED_U_PRODUCTS:
            continue
        exp = ws.cell(9, col).value
        act = u_variant_qty(by_code, u_prod, u_var)
        if act is None:
            if exp not in (None, '', 0):
                mismatches.append(('U_DIRECT_EXCLUDED', '9', f'C{col}', act, exp, 'P00000YU/-'))
            continue
        act = int(act or 0)
        if exp is None:
            if act != 0:
                mismatches.append(('U_DIRECT_BLANK', '9', f'C{col}', act, exp, 'P00000YU/-'))
        else:
            if int(exp) != act:
                mismatches.append(('U_DIRECT_VAL', '9', f'C{col}', act, int(exp), 'P00000YU/-'))

    expected_cache = {}
    for row in rows:
        if row['kind'] != 'data':
            continue
        expected_cache[(row['row'], row['code'], row['variant'])] = row_expected(by_code, row)

    # product/variant rows
    for row in rows:
        if row['kind'] != 'data':
            continue
        code = row['code']
        if code not in L_CODES:
            continue

        act_direct, act_map = expected_cache[(row['row'], row['code'], row['variant'])]
        exp_direct = row['direct']
        if exp_direct != act_direct:
            mismatches.append((
                'DIRECT',
                row['row'],
                'D',
                act_direct,
                exp_direct,
                f"{row['code']}/{row['variant'] or '-'}",
            ))

        for col, idx in sorted(col_map.items()):
            exp = row['cols'].get(col)
            act = act_map.get(idx, 0)
            if exp is None:
                if act != 0:
                    mismatches.append((
                        'MAPPED_BLANK',
                        row['row'],
                        f'C{col}',
                        act,
                        exp,
                        f"{row['code']}/{row['variant'] or '-'}",
                    ))
            else:
                if int(exp) != act:
                    mismatches.append((
                        'MAPPED_VAL',
                        row['row'],
                        f'C{col}',
                        act,
                        int(exp),
                        f"{row['code']}/{row['variant'] or '-'}",
                    ))

    subtotal_map = {r['label']: r for r in rows if r['kind'] == 'subtotal'}

    for group_name, group_codes, with_subtotal in L_GROUPS:
        if not with_subtotal:
            continue
        label = '500g총합계' if group_name == '500g' else '1kg총합계'
        sum_row = subtotal_map.get(label)
        if not sum_row:
            continue

        mapping_codes = set(group_codes)
        if group_name == '1kg':
            mapping_codes.add('P00000ZB')

        exp_direct = int(sum_row['direct'] or 0)
        act_direct = 0
        act_cols = {c: 0 for c in range(5, 59)}

        for row in rows:
            if row['kind'] != 'data':
                continue
            rdir, rmap = expected_cache[(row['row'], row['code'], row['variant'])]
            if row['code'] in group_codes:
                act_direct += rdir
            if row['code'] in mapping_codes:
                for col, idx in col_map.items():
                    act_cols[col] += rmap[idx]

        if act_direct != exp_direct:
            mismatches.append(('SUBTOTAL_DIRECT', label, 'D', act_direct, exp_direct, label))

        for col, idx in sorted(col_map.items()):
            exp = sum_row['cols'].get(col)
            act = act_cols[col]
            if exp is None:
                if act != 0:
                    mismatches.append(('SUBTOTAL_BLANK', label, f'C{col}', act, exp, label))
            else:
                if int(exp) != act:
                    mismatches.append(('SUBTOTAL_VAL', label, f'C{col}', act, int(exp), label))

    print(f'rows_count={len(rows)}')
    print(f'columns_mapped={len(col_map)}/{len(U_COLUMNS)}')
    print(f'mismatch_count={len(mismatches)}')

    for kind, row, col, act, exp, ident in mismatches:
        if row in ('9', '500g총합계', '1kg총합계', '합계', '2kg총합계'):
            if col.startswith('C') and len(col) > 1 and col[1:].isdigit():
                col_idx = int(col[1:])
                row_num = '9' if row == '9' else ('28' if row == '500g총합계' else '47')
                cell = f'{excel_col(col_idx)}{row_num}'
            else:
                cell = f'{col}{row}'
        else:
            if col.startswith('C') and len(col) > 1 and col[1:].isdigit():
                col_idx = int(col[1:])
                cell = f'{excel_col(col_idx)}{row}'
            else:
                cell = f'{col}{row}'
        print(kind, cell, ident, 'act=', act, 'exp=', exp)


if __name__ == '__main__':
    compare()
