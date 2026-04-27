"""
캐치업코리아 cafe24 판매 집계 웹.

실행:
    py web/server.py
브라우저:
    http://127.0.0.1:8000
"""
import sys
import json
import io
import traceback
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "python" / "cafe24" / "tests"))

import requests
from fastapi import FastAPI
from fastapi.responses import StreamingResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from cafe24_auth import get_access_token, MALL_ID
from aggregation import (
    aggregate,
    sort_groups,
    sort_variants,
    sort_categories,
    _group_sort_key,
)

BASE = f"https://{MALL_ID}.cafe24api.com/api/v2/admin"
WEB_DIR = Path(__file__).resolve().parent
INVALID_SHEET = '\\/?*[]:'

app = FastAPI(title="catchup sales report")
app.mount("/static", StaticFiles(directory=WEB_DIR / "static"), name="static")


def auth_headers():
    return {"Authorization": f"Bearer {get_access_token()}"}


def fetch_categories():
    res = requests.get(f"{BASE}/categories", headers=auth_headers(), params={"limit": 100})
    res.raise_for_status()
    return res.json().get("categories", [])


def fetch_products_by_category(cat_no):
    products = {}
    offset = 0
    while True:
        res = requests.get(
            f"{BASE}/products", headers=auth_headers(),
            params={"category": cat_no, "limit": 100, "offset": offset, "embed": "variants"},
        )
        res.raise_for_status()
        items = res.json().get("products", [])
        if not items:
            break
        for p in items:
            vs = []
            for v in (p.get("variants") or []):
                opts = v.get("options") or []
                opt_str = ", ".join(
                    f"{o.get('name')}={o.get('value')}"
                    for o in opts if o.get("value")
                ) if opts else ""
                vs.append({"vcode": v.get("variant_code"), "opt": opt_str})
            products[p["product_no"]] = {
                "code": p["product_code"],
                "name": p["product_name"],
                "price": float(p.get("price") or 0),
                "variants": vs,
            }
        if len(items) < 100:
            break
        offset += 100
    return products


def fetch_orders(start, end):
    orders = []
    offset = 0
    while True:
        res = requests.get(
            f"{BASE}/orders", headers=auth_headers(),
            params={
                "start_date": start, "end_date": end,
                "embed": "items", "limit": 100, "offset": offset,
            },
        )
        res.raise_for_status()
        page = res.json().get("orders", [])
        if not page:
            break
        orders.extend(page)
        if len(page) < 100:
            break
        offset += 100
    return orders


def sanitize_sheet_name(name, used):
    s = name
    for ch in INVALID_SHEET:
        s = s.replace(ch, "_")
    s = s.strip()[:31] or "sheet"
    base = s
    i = 2
    while s in used:
        suf = f"_{i}"
        s = base[: 31 - len(suf)] + suf
        i += 1
    used.add(s)
    return s


def size_cols(ws):
    for col_idx in range(1, ws.max_column + 1):
        col = get_column_letter(col_idx)
        max_len = max((len(str(c.value or "")) for c in ws[col]), default=10)
        ws.column_dimensions[col].width = min(max_len + 2, 60)


CURRENCY_FMT = '"₩"#,##0'
QTY_FMT = '#,##0'


def apply_section_formats(ws):
    """5-column layout: 코드/상품명/단가/판매수/매출"""
    widths = {"A": 14, "B": 60, "C": 13, "D": 10, "E": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    data_font = Font(size=10)
    for r in range(1, ws.max_row + 1):
        a1 = ws.cell(r, 1).value
        is_header = (a1 == "코드") or (a1 == "합계") or (isinstance(a1, str) and a1.startswith("["))
        if not is_header:
            for col in range(1, 6):
                if ws.cell(r, col).font.bold:
                    continue  # parent row의 bold 유지
                ws.cell(r, col).font = data_font
        ws.cell(r, 3).number_format = CURRENCY_FMT
        ws.cell(r, 4).number_format = QTY_FMT
        ws.cell(r, 5).number_format = CURRENCY_FMT


def apply_flat_formats(ws):
    """6-column layout: 카테고리/코드/상품명/단가/판매수/매출"""
    widths = {"A": 30, "B": 14, "C": 60, "D": 13, "E": 10, "F": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    data_font = Font(size=10)
    for r in range(1, ws.max_row + 1):
        a1 = ws.cell(r, 1).value
        is_header = (a1 == "카테고리") or (a1 == "합계")
        if not is_header:
            for col in range(1, 7):
                if ws.cell(r, col).font.bold:
                    continue
                ws.cell(r, col).font = data_font
        ws.cell(r, 4).number_format = CURRENCY_FMT
        ws.cell(r, 5).number_format = QTY_FMT
        ws.cell(r, 6).number_format = CURRENCY_FMT


def parse_categories(s, all_cats):
    if s == "all":
        return [c for c in all_cats if c.get("category_depth") == 1]
    wanted = {int(x.strip()) for x in s.split(",") if x.strip()}
    return [c for c in all_cats if c.get("category_no") in wanted]


def detect_currency(orders):
    for o in orders:
        if o.get("currency"):
            return o["currency"]
    return "KRW"


def sse(data):
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


@app.get("/")
def index():
    return FileResponse(WEB_DIR / "static" / "index.html")


@app.get("/api/categories")
def api_categories():
    try:
        cats = fetch_categories()
        return [
            {
                "no": c["category_no"],
                "name": c["category_name"],
                "depth": c.get("category_depth", 1),
                "parent": c.get("parent_category_no"),
            }
            for c in cats
        ]
    except Exception as e:
        return JSONResponse(
            {"error": f"{type(e).__name__}: {e}", "trace": traceback.format_exc()},
            status_code=500,
        )


@app.get("/api/report")
def api_report(start: str, end: str, categories: str = "all"):
    def gen():
        try:
            yield sse({"type": "progress", "msg": "[1/4] 카테고리 목록 조회"})
            cats = fetch_categories()
            yield sse({"type": "progress", "msg": f"  → {len(cats)}개"})
            target = parse_categories(categories, cats)
            yield sse({"type": "progress", "msg": f"  → 처리 대상: {len(target)}개"})

            yield sse({"type": "progress", "msg": f"[2/4] 주문 조회 ({start} ~ {end})"})
            orders = []
            offset = 0
            while True:
                res = requests.get(f"{BASE}/orders", headers=auth_headers(), params={
                    "start_date": start, "end_date": end,
                    "embed": "items", "limit": 100, "offset": offset,
                })
                res.raise_for_status()
                page = res.json().get("orders", [])
                if not page:
                    break
                orders.extend(page)
                yield sse({"type": "progress", "msg": f"    ... {len(orders)}건 누적"})
                if len(page) < 100:
                    break
                offset += 100
            currency = detect_currency(orders)
            yield sse({"type": "progress", "msg": f"  → 총 {len(orders)}건 / 통화 {currency}"})

            yield sse({"type": "progress", "msg": "[3/4] 카테고리별 상품 + 집계"})
            results = []
            grand_qty = 0
            grand_rev = 0.0
            for i, c in enumerate(target, 1):
                cn = c["category_no"]
                cname = c["category_name"]
                yield sse({"type": "progress", "msg": f"  ({i}/{len(target)}) [{cn}] {cname}"})
                products = fetch_products_by_category(cn)
                groups = aggregate(products, orders)
                cqty = sum(g["qty"] for g in groups)
                crev = sum(g["rev"] for g in groups)
                grand_qty += cqty
                grand_rev += crev
                multi_count = sum(1 for g in groups if g["is_multi"])
                results.append({
                    "category_no": cn,
                    "category_name": cname,
                    "groups": groups,
                    "qty": cqty,
                    "rev": crev,
                })
                yield sse({"type": "progress", "msg": f"    → 상품 {len(groups)} (옵션상품 {multi_count}) / 판매수 {cqty} / 매출 {crev:,.0f}"})

            yield sse({"type": "progress", "msg": "[4/4] 완료"})
            yield sse({
                "type": "data",
                "results": results,
                "grand": {"qty": grand_qty, "rev": grand_rev, "currency": currency, "order_count": len(orders)},
                "start": start,
                "end": end,
            })
            yield sse({"type": "done"})
        except Exception as e:
            yield sse({
                "type": "error",
                "msg": f"{type(e).__name__}: {e}",
                "trace": traceback.format_exc(),
            })

    return StreamingResponse(
        gen(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/api/excel")
def api_excel(
    start: str,
    end: str,
    categories: str = "all",
    mode: str = "single",
    sort_by: str = "rev",
    sort_dir: int = -1,
    cat_sort_by: str = "rev",
    cat_sort_dir: int = -1,
):
    try:
        from openpyxl.styles import Alignment
        from aggregation import _group_sort_key

        all_cats = fetch_categories()
        target = parse_categories(categories, all_cats)
        orders = fetch_orders(start, end)
        currency = detect_currency(orders)

        results = []
        for c in target:
            products = fetch_products_by_category(c["category_no"])
            groups = aggregate(products, orders)
            results.append({
                "category_no": c["category_no"],
                "category_name": c["category_name"],
                "groups": groups,
                "qty": sum(g["qty"] for g in groups),
                "rev": sum(g["rev"] for g in groups),
            })

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        bold = Font(bold=True)
        col_header_font = Font(bold=True, color="FFF8FAFC")
        col_header_fill = PatternFill(start_color="FF334155", end_color="FF334155", fill_type="solid")
        cat_header_font = Font(bold=True, color="FFF8FAFC", size=13)
        cat_header_fill = PatternFill(start_color="FF1E40AF", end_color="FF1E40AF", fill_type="solid")
        sum_fill = PatternFill(start_color="FFFEF3C7", end_color="FFFEF3C7", fill_type="solid")
        parent_fill = PatternFill(start_color="FFF1F5F9", end_color="FFF1F5F9", fill_type="solid")
        center_left = Alignment(vertical="center", horizontal="left")

        def emit_section(ws, cat_label, groups):
            ws.append([cat_label])
            cat_row = ws.max_row
            ws.cell(cat_row, 1).font = cat_header_font
            ws.cell(cat_row, 1).fill = cat_header_fill
            ws.cell(cat_row, 1).alignment = center_left
            ws.row_dimensions[cat_row].height = 24
            ws.merge_cells(start_row=cat_row, start_column=1, end_row=cat_row, end_column=5)

            ws.append(["코드", "상품명", "단가", "판매수", "매출"])
            for c in ws[ws.max_row]:
                c.font = col_header_font
                c.fill = col_header_fill

            cat_qty = 0
            cat_rev = 0
            for g in sort_groups(groups, sort_by, sort_dir):
                cat_qty += g["qty"]
                cat_rev += g["rev"]
                if g["is_multi"]:
                    parent_name = f"{g['product_name']} ({len(g['variants'])}개 옵션)"
                    ws.append([g["product_code"], parent_name, g["price"], g["qty"], g["rev"]])
                    p_row = ws.max_row
                    for c in ws[p_row]:
                        c.fill = parent_fill
                        c.font = bold
                    for v in sort_variants(g["variants"], g["price"], sort_by, sort_dir):
                        suffix = v["variant_code"][len(g["product_code"]):] if v["variant_code"] and v["variant_code"].startswith(g["product_code"]) else (v["variant_code"] or "")
                        label = "  └ " + (v["option"] or v["variant_code"])
                        ws.append([suffix, label, g["price"], v["qty"], v["rev"]])
                        ws.row_dimensions[ws.max_row].outline_level = 1
                else:
                    ws.append([g["product_code"], g["product_name"], g["price"], g["qty"], g["rev"]])

            ws.append(["합계", None, None, cat_qty, cat_rev])
            sum_row = ws.max_row
            for c in ws[sum_row]:
                c.font = bold
                c.fill = sum_fill
            ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
            ws.cell(sum_row, 1).alignment = center_left
            ws.append([])

        if mode == "tabs":
            sorted_results = sort_categories(results, cat_sort_by, cat_sort_dir)
            used = set()
            for r in sorted_results:
                title = sanitize_sheet_name(f"{r['category_no']}_{r['category_name']}", used)
                ws = wb.create_sheet(title=title)
                ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
                ws.append([])
                cat_label = f"[{r['category_no']}] {r['category_name']}"
                emit_section(ws, cat_label, r["groups"])
                ws.freeze_panes = "A4"
                apply_section_formats(ws)
        elif mode == "flat":
            ws = wb.create_sheet(title="전체")
            ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
            ws.append([])
            ws.append(["카테고리", "코드", "상품명", "단가", "판매수", "매출"])
            for c in ws[ws.max_row]:
                c.font = col_header_font
                c.fill = col_header_fill

            all_groups = []
            grand_q = grand_r = 0
            for r in results:
                cat_label = f"[{r['category_no']}] {r['category_name']}"
                for g in r["groups"]:
                    all_groups.append((cat_label, g))
                    grand_q += g["qty"]
                    grand_r += g["rev"]
            all_groups.sort(key=lambda x: _group_sort_key(x[1], sort_by), reverse=(sort_dir == -1))

            for cat_label, g in all_groups:
                if g["is_multi"]:
                    parent_name = f"{g['product_name']} ({len(g['variants'])}개 옵션)"
                    ws.append([cat_label, g["product_code"], parent_name, g["price"], g["qty"], g["rev"]])
                    p_row = ws.max_row
                    for c in ws[p_row]:
                        c.fill = parent_fill
                        c.font = bold
                    for v in sort_variants(g["variants"], g["price"], sort_by, sort_dir):
                        suffix = v["variant_code"][len(g["product_code"]):] if v["variant_code"] and v["variant_code"].startswith(g["product_code"]) else (v["variant_code"] or "")
                        label = "  └ " + (v["option"] or v["variant_code"])
                        ws.append([cat_label, suffix, label, g["price"], v["qty"], v["rev"]])
                        ws.row_dimensions[ws.max_row].outline_level = 1
                else:
                    ws.append([cat_label, g["product_code"], g["product_name"], g["price"], g["qty"], g["rev"]])

            ws.append(["합계", None, None, None, grand_q, grand_r])
            sum_row = ws.max_row
            for c in ws[sum_row]:
                c.font = bold
                c.fill = sum_fill
            ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4)
            ws.cell(sum_row, 1).alignment = center_left
            ws.freeze_panes = "A4"
            apply_flat_formats(ws)
        else:
            ws = wb.create_sheet(title="합산")
            ws.append([f"기간: {start} ~ {end} / 통화: {currency}"])
            ws.append([])
            sorted_results = sort_categories(results, cat_sort_by, cat_sort_dir)
            for r in sorted_results:
                cat_label = f"[{r['category_no']}] {r['category_name']}"
                emit_section(ws, cat_label, r["groups"])
            ws.freeze_panes = "A3"
            apply_section_formats(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fn = f"sales_report_{start}_{end}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fn}"'},
        )
    except Exception as e:
        return JSONResponse(
            {"error": f"{type(e).__name__}: {e}", "trace": traceback.format_exc()},
            status_code=500,
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
